# app.py
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from st_aggrid import AgGrid, GridOptionsBuilder
from datetime import datetime
import textwrap
from pathlib import Path
from openpyxl import load_workbook

# ---------------- Configurações ----------------
EXCEL_PATH = "Acompto_Abast.xlsx"

# Paletas
PALETTE_LIGHT = px.colors.sequential.Blues_r
PALETTE_DARK = px.colors.sequential.Plasma_r

# Classes que serão agrupadas em "Outros"
OUTROS_CLASSES = {"Motocicletas", "Mini Carregadeira", "Usina", "Veiculos Leves"}

# Nome exato (opcional) da coluna única que conterá hodômetro / horímetro atual na aba BD
COL_KM_HR_ATUAL = "KM_HR_Atual"  # se criada na BD, será priorizada

# Nome da aba de log de manutenção
MANUT_LOG_SHEET = "MANUTENCAO_LOG"

# ---------------- Utilitários ----------------
def formatar_brasileiro(valor):
    if pd.isna(valor) or not np.isfinite(valor):
        return "–"
    return "{:,.2f}".format(valor).replace(",", "X").replace(".", ",").replace("X", ".")

def wrap_labels(s: str, width: int = 18) -> str:
    if pd.isna(s):
        return ""
    parts = textwrap.wrap(str(s), width=width)
    return "<br>".join(parts) if parts else str(s)

def find_first_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for c in candidates:
        if c in df.columns:
            return c
    return None

# candidatos comuns
CANDIDATES_COD = ["Cod_Equip", "COD_EQUIPAMENTO", "Cod Equip", "CODIGO", "COD", "Código", "Code"]
CANDIDATES_UNID = ["Unid", "Unidade", "UNID", "UNIDADE"]
CANDIDATES_KMHS = ["Km_Hs_Rod", "KM_HS_ROD", "KM_HS", "KM", "Km", "Km_Hs"]
CANDIDATES_LITROS = ["Qtde_Litros", "Litros", "Qtd_Litros", "QTDE_LITROS"]
CANDIDATES_DESCR = ["Descricao_Equip", "DESCRICAO_EQUIPAMENTO", "Descricao", "DESCRICAO"]
CANDIDATES_FROTA_COD = ["COD_EQUIPAMENTO", "Cod_Equip", "Cod_Equipamento"]

# ---------------- Carregamento & Normalização ----------------
@st.cache_data(show_spinner="Carregando e normalizando dados...")
def load_data(path: str):
    """
    Lê o arquivo Excel (BD e FROTAS), detecta colunas relevantes mesmo com nomes diferentes
    e cria colunas consolidadas usadas pelo app:
      - Cod_Equip (código do equipamento)
      - Descricao_Equip
      - Qtde_Litros
      - Km_Hs_Rod
      - Unidade / Unid
      - Valor_Atual (KM_HR_Atual ou fallback pelo último Km_Hs_Rod)
      - Media (km/l ou hr/l) calculada automaticamente com base em Unid e valores disponíveis
    """
    if not Path(path).exists():
        st.error(f"Arquivo não encontrado: {path}")
        st.stop()

    # tenta ler ambas abas (se existirem)
    try:
        df_bd = pd.read_excel(path, sheet_name="BD", skiprows=2)
    except Exception as e:
        st.error(f"Erro ao ler sheet 'BD': {e}")
        st.stop()
    try:
        df_frotas = pd.read_excel(path, sheet_name="FROTAS", skiprows=1)
    except Exception as e:
        st.error(f"Erro ao ler sheet 'FROTAS': {e}")
        st.stop()

    # detectar coluna de código do equipamento em BD e FROTAS
    cod_bd = find_first_column(df_bd, CANDIDATES_COD)
    cod_frotas = find_first_column(df_frotas, CANDIDATES_FROTA_COD) or find_first_column(df_frotas, CANDIDATES_COD)

    # normalizar nomes mínimos em BD
    # cria cópias para evitar alterar original
    df_abast = df_bd.copy()
    df_f = df_frotas.copy()

    # padroniza coluna de código em ambas as tabelas
    if cod_bd:
        df_abast = df_abast.rename(columns={cod_bd: "Cod_Equip"})
    else:
        # se não encontrou, tenta buscar coluna similar 'Cod_Equip' direto
        if "Cod_Equip" not in df_abast.columns:
            # sem código, será impossível mapear histórico por equipamento; manter BD mas alertar
            st.warning("Coluna com Código do Equipamento não encontrada na aba 'BD'. Alguns recursos por equipamento ficarão limitados.")
    if cod_frotas:
        df_f = df_f.rename(columns={cod_frotas: "Cod_Equip"})
    else:
        st.error("Coluna de código do equipamento não encontrada na aba 'FROTAS'. Verifique planilha.")
        st.stop()

    # detecta outras colunas com nomes variados
    kmhs_col = find_first_column(df_abast, CANDIDATES_KMHS)
    litros_col = find_first_column(df_abast, CANDIDATES_LITROS)
    descr_col = find_first_column(df_abast, CANDIDATES_DESCR)
    unid_col = find_first_column(df_abast, CANDIDATES_UNID)

    # renomeia para padrão onde existir
    col_map = {}
    if kmhs_col:
        col_map[kmhs_col] = "Km_Hs_Rod"
    if litros_col:
        col_map[litros_col] = "Qtde_Litros"
    if descr_col:
        col_map[descr_col] = "Descricao_Equip"
    if unid_col:
        col_map[unid_col] = "Unidade"
    if col_map:
        df_abast = df_abast.rename(columns=col_map)

    # se houver coluna KM_HR_ATUAL na BD, priorizamos (pode ser criada pelo usuário)
    if COL_KM_HR_ATUAL in df_abast.columns:
        df_abast[COL_KM_HR_ATUAL] = pd.to_numeric(df_abast[COL_KM_HR_ATUAL], errors="coerce")

    # padroniza colunas em df_frotas
    if "DESCRICAO_EQUIPAMENTO" in df_f.columns and "DESCRICAO_EQUIPAMENTO" not in df_f.columns:
        pass  # safety
    # tentar normalizar descrição e placa
    if "DESCRICAO_EQUIPAMENTO" in df_f.columns and "DESCRICAO_EQUIPAMENTO" not in df_f.columns:
        pass

    # garantir tipos
    df_abast["Data"] = pd.to_datetime(df_abast["Data"], errors="coerce")
    df_abast = df_abast.dropna(subset=["Data"])

    # preenche colunas mínimas caso não existam (para não quebrar o restante)
    for c in ["Cod_Equip", "Descricao_Equip", "Qtde_Litros", "Km_Hs_Rod", "Unidade"]:
        if c not in df_abast.columns:
            df_abast[c] = np.nan

    # normaliza df_frotas: renomear COD_EQUIPAMENTO se existir
    if "COD_EQUIPAMENTO" in df_f.columns and "Cod_Equip" not in df_f.columns:
        df_f = df_f.rename(columns={"COD_EQUIPAMENTO": "Cod_Equip"})
    if "Cod_Equip" not in df_f.columns:
        st.error("Planilha FROTAS precisa ter a coluna de código do equipamento (ex.: 'COD_EQUIPAMENTO' ou 'Cod_Equip').")
        st.stop()

    # tenta converter ANOMODELO
    if "ANOMODELO" in df_f.columns:
        df_f["ANOMODELO"] = pd.to_numeric(df_f["ANOMODELO"], errors="coerce")
    else:
        df_f["ANOMODELO"] = np.nan

    # cria label amigável em frotas
    placa_col = None
    for cand in ["PLACA", "Placa", "placa"]:
        if cand in df_f.columns:
            placa_col = cand
            break
    desc_frota_col = None
    for cand in ["DESCRICAO_EQUIPAMENTO", "DESCRICAO", "Descricao_Equip", "Descricao"]:
        if cand in df_f.columns:
            desc_frota_col = cand
            break
    df_f["label"] = df_f["Cod_Equip"].astype(str) + " - " + df_f.get(desc_frota_col, "").fillna("").astype(str) + " (" + df_f.get(placa_col, "").fillna("Sem Placa").astype(str) + ")"

    # merge histórico com frotas para enriquecer
    # alguns arquivos têm Cod_Equip como inteiro, outros como str -> cast para str para merge seguro
    df_abast["Cod_Equip"] = df_abast["Cod_Equip"].astype(str).str.strip()
    df_f["Cod_Equip"] = df_f["Cod_Equip"].astype(str).str.strip()

    df = pd.merge(df_abast, df_f, on="Cod_Equip", how="left", suffixes=("_bd", "_frota"))

    # campos derivados
    df["Mes"] = df["Data"].dt.month
    df["Semana"] = df["Data"].dt.isocalendar().week
    df["Ano"] = df["Data"].dt.year
    df["AnoMes"] = df["Data"].dt.to_period("M").astype(str)
    df["AnoSemana"] = df["Data"].dt.strftime("%Y-%U")

    # converte numéricos importantes
    for col in ["Qtde_Litros", "Km_Hs_Rod"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # cria Valor_Atual: prioriza COL_KM_HR_ATUAL (se existir) — mas atenção: COL_KM_HR_ATUAL pode estar em BD ou FROTAS
    valor_atual = None
    if COL_KM_HR_ATUAL in df.columns:
        valor_atual = pd.to_numeric(df[COL_KM_HR_ATUAL], errors="coerce")
        df["Valor_Atual"] = valor_atual
    else:
        # fallback: último Km_Hs_Rod por equipamento se disponível
        last_km = df.sort_values(["Cod_Equip", "Data"]).groupby("Cod_Equip")["Km_Hs_Rod"].last().rename("Last_Km")
        # merge last_km to rows (so each row has valor, but we'll also use groupby last per equipment later)
        df = df.merge(last_km, how="left", left_on="Cod_Equip", right_index=True)
        df["Valor_Atual"] = df["Last_Km"]

    # detectar coluna de unidade
    if "Unidade" not in df.columns:
        # tentar variações
        for cand in ["Unid", "Unidade", "UNID", "UNIDADE"]:
            if cand in df.columns:
                df = df.rename(columns={cand: "Unidade"})
                break
        if "Unidade" not in df.columns:
            # deixar coluna vazia para evitar erros
            df["Unidade"] = np.nan

    # cálculo automático de Media baseado na Unidade:
    # - se Unid contém 'QUIL' ou 'KM' -> usar Km_Hs_Rod / Qtde_Litros (km/l)
    # - se Unid contém 'HOR' ou 'HR' -> usar Km_Hs_Rod (que representa horas) / Qtde_Litros (hr/l)
    # - se Qtde_Litros <=0 => NaN
    def calc_media(row):
        litros = row.get("Qtde_Litros", np.nan)
        kmhs = row.get("Km_Hs_Rod", np.nan)
        un = str(row.get("Unidade", "")).upper() if pd.notna(row.get("Unidade")) else ""
        if pd.isna(litros) or litros == 0:
            return np.nan
        if "QUIL" in un or "KM" in un:
            return kmhs / litros if not pd.isna(kmhs) else np.nan
        if "HOR" in un or "HR" in un:
            # tratar Horas como 'kmhs' aqui (horas run) -> horas por litro
            return kmhs / litros if not pd.isna(kmhs) else np.nan
        # fallback: tentar usar Km_Hs_Rod / litros
        return kmhs / litros if not pd.isna(kmhs) else np.nan

    # aplica cálculo por linha
    df["Media_calc"] = df.apply(calc_media, axis=1)
    # Se existir coluna 'Media' já fornecida, mantemos, senão substituímos por calc
    if "Media" in df.columns:
        # se Media tem muitos NaNs, preferir a calculada; caso contrário manter original mas preencher NaNs com calculada
        na_ratio = df["Media"].isna().mean()
        if na_ratio > 0.3:
            df["Media"] = df["Media_calc"]
        else:
            df["Media"] = df["Media"].fillna(df["Media_calc"])
    else:
        df["Media"] = df["Media_calc"]

    # para segurança: converter Media para numérico
    df["Media"] = pd.to_numeric(df["Media"], errors="coerce")

    # formato final: garantir df_f (frotas) tem Cod_Equip e colunas básicas
    # converte tipos
    df_f["Cod_Equip"] = df_f["Cod_Equip"].astype(str)
    df["Cod_Equip"] = df["Cod_Equip"].astype(str)

    return df, df_f

# ---------------- Salvar log de manutenção ----------------
def save_maintenance_log(excel_path: str, entries_df: pd.DataFrame, sheet_name: str = MANUT_LOG_SHEET) -> bool:
    """
    Salva (append) as entradas de manutenção em uma aba MANUTENCAO_LOG.
    Se a aba existir, concatena; senão cria.
    """
    try:
        # formatos
        if "Timestamp" in entries_df.columns:
            entries_df["Timestamp"] = pd.to_datetime(entries_df["Timestamp"])

        # se arquivo não existir, cria com o sheet
        if not Path(excel_path).exists():
            with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
                entries_df.to_excel(writer, sheet_name=sheet_name, index=False)
            return True

        # se existir, ler a planilha existente de log (se houver) e concatenar
        book = load_workbook(excel_path)
        if sheet_name in book.sheetnames:
            existing = pd.read_excel(excel_path, sheet_name=sheet_name)
            combined = pd.concat([existing, entries_df], ignore_index=True)
        else:
            combined = entries_df

        # reescrever (substituir) a sheet de log
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            combined.to_excel(writer, sheet_name=sheet_name, index=False)
        return True
    except Exception as e:
        st.error(f"Erro ao gravar log de manutenção: {e}")
        return False

# ---------------- Manutenção: construção do plano ----------------
def build_maintenance_dataframe(df_frotas, df_abast, class_intervals, km_default, hr_default):
    """
    Constrói tabela com próximos serviços/lubrificação por equipamento.
    - class_intervals: dict[class] = {"rev_km": [r1,r2,r3], "rev_hr": [r1,r2,r3]}
    """
    mf = df_frotas.copy()
    mf["Cod_Equip"] = mf["Cod_Equip"].astype(str)

    # pegar último Valor_Atual por equipamento (do df_abast já enriquecido)
    try:
        last_vals = df_abast.sort_values(["Cod_Equip", "Data"]).groupby("Cod_Equip")["Valor_Atual"].last()
    except Exception:
        last_vals = pd.Series(dtype=float)

    mf = mf.set_index("Cod_Equip")
    mf["Km_Hr_Atual"] = last_vals.reindex(mf.index).astype(float)
    mf = mf.reset_index()

    # pega última Unidade (Unidade/Unid) por equipamento do histórico
    try:
        last_unid = df_abast.sort_values(["Cod_Equip", "Data"]).groupby("Cod_Equip")["Unidade"].last()
    except Exception:
        last_unid = pd.Series(dtype=object)
    mf = mf.set_index("Cod_Equip")
    mf["Unid_Last"] = last_unid.reindex(mf.index)
    mf = mf.reset_index()

    # se frotas tem coluna Unid/Unidade prioriza
    if "Unid" in df_frotas.columns and "Unid" not in mf.columns:
        mf["Unid"] = df_frotas.set_index("Cod_Equip").reindex(mf["Cod_Equip"]).get("Unid").values
    # compor Unid final
    mf["Unid"] = mf.get("Unid", np.nan)
    mf["Unid"] = np.where(pd.isna(mf["Unid"]), mf["Unid_Last"], mf["Unid"])
    mf["Unid"] = mf["Unid"].fillna("")

    revs = [1, 2, 3]
    for r in revs:
        # valor por classe se disponível, senão defaults
        def get_km_interval(cls):
            if cls in class_intervals and "rev_km" in class_intervals[cls] and len(class_intervals[cls]["rev_km"]) >= r:
                return class_intervals[cls]["rev_km"][r-1]
            return km_default * r

        def get_hr_interval(cls):
            if cls in class_intervals and "rev_hr" in class_intervals[cls] and len(class_intervals[cls]["rev_hr"]) >= r:
                return class_intervals[cls]["rev_hr"][r-1]
            return hr_default * r

        mf[f"Rev{r}_Interval_KM"] = mf["Classe_Operacional"].apply(lambda c: get_km_interval(c if pd.notna(c) else ""))
        mf[f"Rev{r}_Interval_HR"] = mf["Classe_Operacional"].apply(lambda c: get_hr_interval(c if pd.notna(c) else ""))

    # calcular next e to_go baseado em Unid (QUILÔMETROS vs HORAS)
    def compute_next(row, r):
        cur = row.get("Km_Hr_Atual", np.nan)
        unit = str(row.get("Unid", "")).upper() if pd.notna(row.get("Unid")) else ""
        if pd.isna(cur):
            return (np.nan, np.nan)
        if "QUIL" in unit or "KM" in unit:
            interval = row.get(f"Rev{r}_Interval_KM")
            next_due = cur + (interval if not pd.isna(interval) else np.nan)
            to_go = next_due - cur
            return (next_due, to_go)
        if "HOR" in unit or "HR" in unit:
            interval = row.get(f"Rev{r}_Interval_HR")
            next_due = cur + (interval if not pd.isna(interval) else np.nan)
            to_go = next_due - cur
            return (next_due, to_go)
        # fallback: assume KM
        interval = row.get(f"Rev{r}_Interval_KM")
        next_due = cur + (interval if not pd.isna(interval) else np.nan)
        to_go = next_due - cur
        return (next_due, to_go)

    for r in revs:
        mf[[f"Rev{r}_Next", f"Rev{r}_To_Go"]] = mf.apply(lambda row: pd.Series(compute_next(row, r)), axis=1)

    # flags placeholders
    mf["Due_Rev"] = False
    mf["Due_Oil"] = False
    mf["Any_Due"] = False

    return mf

# ---------------- Visual / CSS ----------------
def apply_modern_css(dark: bool):
    bg = "#0e1117" if dark else "#FFFFFF"
    card_bg = "#111318" if dark else "#f8f9fa"
    text_color = "#f0f0f0" if dark else "#111111"
    st.markdown(
        f"""
        <style>
        .stApp {{
            background-color: {bg};
            color: {text_color};
        }}
        .kpi-card {{
            background: {card_bg};
            padding: 12px;
            border-radius: 10px;
            box-shadow: 0 2px 6px rgba(0,0,0,0.08);
        }}
        .kpi-title {{ font-size:14px; color: {text_color}; opacity:0.9 }}
        .kpi-value {{ font-size:20px; font-weight:700; color: {text_color} }}
        </style>
        """,
        unsafe_allow_html=True
    )

def make_bar(fig_df, x, y, title, labels, palette, rotate_x=-45, ticksize=10, height=None, hoverfmt=None, wrap_width=16, hide_text_if_gt=10):
    df_local = fig_df.copy()
    if x in df_local.columns:
        df_local[x] = df_local[x].astype(str).apply(lambda s: wrap_labels(s, width=wrap_width))
    fig = px.bar(df_local, x=x, y=y, text=y, title=title, labels=labels, color_discrete_sequence=palette)
    if df_local.shape[0] > hide_text_if_gt:
        fig.update_traces(texttemplate=None)
    else:
        fig.update_traces(texttemplate="%{text:.1f}", textfont=dict(size=10))
    fig.update_layout(
        xaxis=dict(tickangle=rotate_x, tickfont=dict(size=ticksize), automargin=True),
        margin=dict(l=40, r=20, t=60, b=130),
        title=dict(x=0.01, xanchor="left"),
        font=dict(size=13)
    )
    if height:
        fig.update_layout(height=height)
    if hoverfmt:
        fig.update_traces(hovertemplate=hoverfmt)
    else:
        fig.update_traces(hovertemplate=None)
    return fig

# ---------------- App principal ----------------
def main():
    st.set_page_config(page_title="Dashboard Frotas & Manutenção", layout="wide")
    st.title("📊 Dashboard de Frotas e Abastecimentos — Manutenção integrada")

    # Carrega dados
    df, df_frotas = load_data(EXCEL_PATH)

    # Sidebar
    with st.sidebar:
        st.header("Configurações")
        dark_mode = st.checkbox("🕶️ Dark Mode", value=False)
        st.markdown("---")
        st.header("Visual")
        top_n = st.slider("Top N classes antes de agrupar 'Outros'", min_value=3, max_value=30, value=10)
        hide_text_threshold = st.slider("Esconder valores nas barras quando categorias >", min_value=5, max_value=40, value=8)
        st.markdown("---")
        st.header("Manutenção - Defaults")
        km_interval_default = st.number_input("Intervalo padrão (km) revisão", min_value=100, max_value=200000, value=10000, step=100)
        hr_interval_default = st.number_input("Intervalo padrão (horas) lubrificação", min_value=1, max_value=5000, value=250, step=1)
        km_due_threshold = st.number_input("Alerta revisão se faltar <= (km)", min_value=10, max_value=5000, value=500, step=10)
        hr_due_threshold = st.number_input("Alerta lubrificação se faltar <= (horas)", min_value=1, max_value=500, value=20, step=1)
        st.markdown("---")
        if st.button("🔄 Limpar Sessão (reiniciar)"):
            st.session_state.clear()
            st.experimental_rerun()

    apply_modern_css(dark_mode)
    palette = PALETTE_DARK if dark_mode else PALETTE_LIGHT
    plotly_template = "plotly_dark" if dark_mode else "plotly"

    # init session storage for manut_by_class
    if "manut_by_class" not in st.session_state:
        st.session_state.manut_by_class = {}
        classes = sorted(df["Classe_Operacional"].dropna().unique()) if "Classe_Operacional" in df.columns else []
        for cls in classes:
            st.session_state.manut_by_class[cls] = {
                "rev_km": [int(km_interval_default), int(km_interval_default*2), int(km_interval_default*3)],
                "rev_hr": [int(hr_interval_default), int(hr_interval_default*2), int(hr_interval_default*3)]
            }

    tabs = st.tabs(["📊 Análise de Consumo", "🔎 Consulta de Frota", "📋 Tabela Detalhada", "⚙️ Configurações", "🛠️ Manutenção"])

    # ----- Aba Análise -----
    with tabs[0]:
        st.header("Análise de Consumo")
        st.info("Gráficos de média por classe (Media = km/l ou hr/l conforme Unidade).")
        df_plot = df.copy()
        # garantir coluna Classe_Operacional (vários nomes possíveis)
        if "Classe_Operacional" not in df_plot.columns:
            # tentar variações
            for cand in ["Classe Operacional", "Classe_Operacional_bd", "Classe_Operacional_frota", "Classe"]:
                if cand in df_plot.columns:
                    df_plot = df_plot.rename(columns={cand: "Classe_Operacional"})
                    break
        df_plot["Classe_Operacional"] = df_plot.get("Classe_Operacional", "Sem Classe").fillna("Sem Classe")
        df_plot["Classe_Grouped"] = df_plot["Classe_Operacional"].apply(lambda s: "Outros" if s in OUTROS_CLASSES else s)
        media_op_full = df_plot.groupby("Classe_Grouped")["Media"].mean().reset_index()
        media_op_full["Media"] = media_op_full["Media"].round(1)
        media_sorted = media_op_full.sort_values("Media", ascending=False)
        if media_sorted.shape[0] > top_n:
            top_keep = media_sorted.head(top_n)["Classe_Grouped"].tolist()
            df_plot["Classe_TopN"] = df_plot["Classe_Grouped"].apply(lambda s: s if s in top_keep else "Outros")
            media_op = df_plot.groupby("Classe_TopN")["Media"].mean().reset_index().rename(columns={"Classe_TopN":"Classe_Grouped"})
            media_op["Media"] = media_op["Media"].round(1)
        else:
            media_op = media_sorted
        media_op["Classe_wrapped"] = media_op["Classe_Grouped"].apply(lambda s: wrap_labels(s, width=16))
        hover_template_media = "Classe: %{x}<br>Média: %{y:.1f} (km/l ou hr/l)<extra></extra>"
        fig = make_bar(media_op, "Classe_wrapped", "Media", "Média de Consumo por Classe Operacional", {"Media":"Média", "Classe_wrapped":"Classe"}, palette, rotate_x=-60, ticksize=10, height=520, hoverfmt=hover_template_media, wrap_width=16, hide_text_if_gt=hide_text_threshold)
        fig.update_layout(template=plotly_template)
        st.plotly_chart(fig, use_container_width=True, theme=None)

    # ----- Aba Consulta de Frota -----
    with tabs[1]:
        st.header("Ficha Individual do Equipamento")
        equip_opts = df_frotas.sort_values("Cod_Equip")["label"].tolist()
        equip_label = st.selectbox("Selecione o Equipamento", options=equip_opts)
        if equip_label:
            cod_sel = str(equip_label.split(" - ")[0])
            dados_eq = df_frotas.query("Cod_Equip == @cod_sel").iloc[0]
            # último valor atual no histórico/enriquecido
            last_val_series = df.sort_values(["Cod_Equip","Data"]).groupby("Cod_Equip")["Valor_Atual"].last()
            val = last_val_series.get(cod_sel, np.nan)
            un_last = df.sort_values(["Cod_Equip","Data"]).groupby("Cod_Equip")["Unidade"].last().get(cod_sel, "")
            st.subheader(f"{dados_eq.get('DESCRICAO_EQUIPAMENTO', '')} ({dados_eq.get('PLACA','–')})")
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Status", dados_eq.get("ATIVO", "–"))
            c2.metric("Placa", dados_eq.get("PLACA", "–"))
            c3.metric("Medida Atual", f"{formatar_brasileiro(val)} {un_last}")
            consumo_eq = df.query("Cod_Equip == @cod_sel")
            c4.metric("Média Geral", formatar_brasileiro(consumo_eq["Media"].mean()))

    # ----- Aba Tabela -----
    with tabs[2]:
        st.header("Tabela Detalhada de Abastecimentos")
        cols = ["Data", "Cod_Equip", "Descricao_Equip", "PLACA", "DESCRICAOMARCA", "ANOMODELO", "Qtde_Litros", "Km_Hs_Rod", "Media", "Unidade", "Valor_Atual"]
        df_tab = df[[c for c in cols if c in df.columns]]
        st.download_button("⬇️ Exportar CSV da Tabela", df_tab.to_csv(index=False).encode("utf-8"), "abastecimentos.csv", "text/csv")
        gb = GridOptionsBuilder.from_dataframe(df_tab)
        gb.configure_default_column(filterable=True, sortable=True, resizable=True)
        if "Media" in df_tab.columns:
            gb.configure_column("Media", type=["numericColumn"], precision=1)
        if "Qtde_Litros" in df_tab.columns:
            gb.configure_column("Qtde_Litros", type=["numericColumn"], precision=1)
        gb.configure_pagination(paginationAutoPageSize=False, paginationPageSize=15)
        gb.configure_selection("single", use_checkbox=True)
        AgGrid(df_tab, gridOptions=gb.build(), height=520, allow_unsafe_jscode=True)

    # ----- Aba Configurações -----
    with tabs[3]:
        st.header("Padrões por Classe Operacional (Intervalos de Revisão / Horas)")
        classes = sorted(df["Classe_Operacional"].dropna().unique()) if "Classe_Operacional" in df.columns else []
        if "manut_by_class" not in st.session_state:
            st.session_state.manut_by_class = {}
            for cls in classes:
                st.session_state.manut_by_class[cls] = {
                    "rev_km": [int(km_interval_default), int(km_interval_default*2), int(km_interval_default*3)],
                    "rev_hr": [int(hr_interval_default), int(hr_interval_default*2), int(hr_interval_default*3)]
                }

        st.markdown("Ajuste os intervalos por classe (três revisões).")
        for cls in classes:
            st.subheader(str(cls))
            col1, col2 = st.columns(2)
            prev = st.session_state.manut_by_class.get(cls, {})
            rev_km = prev.get("rev_km", [km_interval_default, km_interval_default*2, km_interval_default*3])
            rev_hr = prev.get("rev_hr", [hr_interval_default, hr_interval_default*2, hr_interval_default*3])
            with col1:
                nk1 = st.number_input(f"{cls} → Rev1 (km)", min_value=0, max_value=1000000, value=int(rev_km[0]), key=f"{cls}_r1km")
                nk2 = st.number_input(f"{cls} → Rev2 (km)", min_value=0, max_value=1000000, value=int(rev_km[1]), key=f"{cls}_r2km")
                nk3 = st.number_input(f"{cls} → Rev3 (km)", min_value=0, max_value=1000000, value=int(rev_km[2]), key=f"{cls}_r3km")
            with col2:
                nh1 = st.number_input(f"{cls} → Rev1 (hr)", min_value=0, max_value=1000000, value=int(rev_hr[0]), key=f"{cls}_r1hr")
                nh2 = st.number_input(f"{cls} → Rev2 (hr)", min_value=0, max_value=1000000, value=int(rev_hr[1]), key=f"{cls}_r2hr")
                nh3 = st.number_input(f"{cls} → Rev3 (hr)", min_value=0, max_value=1000000, value=int(rev_hr[2]), key=f"{cls}_r3hr")
            st.session_state.manut_by_class[cls] = {"rev_km": [int(nk1), int(nk2), int(nk3)], "rev_hr": [int(nh1), int(nh2), int(nh3)]}

    # ----- Aba Manutenção -----
    with tabs[4]:
        st.header("Controle de Revisões e Lubrificação")
        st.markdown("O sistema usa `KM_HR_Atual` (se houver) ou o último registro de Km/Hr no histórico (BD). A coluna de unidade deve ser `Unidade` com valores como 'QUILÔMETROS' ou 'HORAS'.")

        # compor class_intervals a partir da sessão
        class_intervals = {}
        for k, v in st.session_state.manut_by_class.items():
            class_intervals[k] = {"rev_km": v.get("rev_km", []), "rev_hr": v.get("rev_hr", [])}

        mf = build_maintenance_dataframe(df_frotas, df, class_intervals, int(km_interval_default), int(hr_interval_default))

        # calcular flags due
        def set_flags(row):
            due_km = False
            due_hr = False
            unit = str(row.get("Unid", "") or row.get("Unid_Last","") or row.get("Unid","")).upper() if pd.notna(row.get("Unid","")) or pd.notna(row.get("Unid_Last","")) else ""
            for r in [1,2,3]:
                to_go = row.get(f"Rev{r}_To_Go", np.nan)
                if pd.isna(to_go):
                    continue
                # avaliar respectiva unidade
                if "QUIL" in unit or "KM" in unit:
                    if to_go <= km_due_threshold:
                        due_km = True
                elif "HOR" in unit or "HR" in unit:
                    if to_go <= hr_due_threshold:
                        due_hr = True
                else:
                    # sem unidade clara -> usar km threshold
                    if to_go <= km_due_threshold:
                        due_km = True
            return pd.Series({"Due_Rev": due_km, "Due_Oil": due_hr})

        if not mf.empty:
            flags = mf.apply(set_flags, axis=1)
            mf["Due_Rev"] = flags["Due_Rev"]
            mf["Due_Oil"] = flags["Due_Oil"]
            mf["Any_Due"] = mf["Due_Rev"] | mf["Due_Oil"]
        else:
            mf["Any_Due"] = False

        df_due = mf[mf["Any_Due"]].copy().sort_values(["Due_Rev", "Due_Oil"], ascending=False)

        st.subheader("Equipamentos com revisão/lubrificação próxima ou vencida")
        st.write(f"Total: {len(df_due)}")
        if not df_due.empty:
            display_cols = ["Cod_Equip", "DESCRICAO_EQUIPAMENTO", "Unid", "Km_Hr_Atual",
                            "Rev1_To_Go", "Rev2_To_Go", "Rev3_To_Go", "Due_Rev", "Due_Oil"]
            available = [c for c in display_cols if c in df_due.columns]
            st.dataframe(df_due[available].reset_index(drop=True), use_container_width=True)

            st.markdown("### Marcar revisões como concluídas")
            st.markdown("Marque as caixas correspondentes e clique em **Salvar ações** para gravar no Excel (aba MANUTENCAO_LOG).")

            actions = []
            for idx, row in df_due.reset_index(drop=True).iterrows():
                cod = row["Cod_Equip"]
                name = row.get("DESCRICAO_EQUIPAMENTO", "")
                unit = row.get("Unid", "")
                cur_val = row.get("Km_Hr_Atual", np.nan)
                st.markdown(f"**{cod} - {name}** — Atual: {formatar_brasileiro(cur_val)} {unit}")
                cols = st.columns([1,1,1,1])
                cb1 = cols[0].checkbox(f"Rev1 (cod {cod})", key=f"r1_{cod}")
                cb2 = cols[1].checkbox(f"Rev2 (cod {cod})", key=f"r2_{cod}")
                cb3 = cols[2].checkbox(f"Rev3 (cod {cod})", key=f"r3_{cod}")
                cbd = cols[3].checkbox(f"Lubrificação (cod {cod})", key=f"lub_{cod}")
                if cb1:
                    actions.append({"Cod_Equip": cod, "Tipo":"Rev1", "Valor_Atual": cur_val, "Unid":unit})
                if cb2:
                    actions.append({"Cod_Equip": cod, "Tipo":"Rev2", "Valor_Atual": cur_val, "Unid":unit})
                if cb3:
                    actions.append({"Cod_Equip": cod, "Tipo":"Rev3", "Valor_Atual": cur_val, "Unid":unit})
                if cbd:
                    actions.append({"Cod_Equip": cod, "Tipo":"Lubrificacao", "Valor_Atual": cur_val, "Unid":unit})

            if st.button("💾 Salvar ações de manutenção"):
                if not actions:
                    st.info("Nenhuma ação selecionada.")
                else:
                    now = datetime.now()
                    rows = []
                    for a in actions:
                        rows.append({
                            "Timestamp": now,
                            "Cod_Equip": a["Cod_Equip"],
                            "Tipo": a["Tipo"],
                            "Valor_Atual": a["Valor_Atual"],
                            "Unid": a["Unid"],
                            "Usuario": st.session_state.get("usuario","(anon)")
                        })
                    entries_df = pd.DataFrame(rows)
                    ok = save_maintenance_log(EXCEL_PATH, entries_df, MANUT_LOG_SHEET)
                    if ok:
                        st.success(f"{len(rows)} ação(ões) registrada(s) em `{MANUT_LOG_SHEET}`.")
                        st.experimental_rerun()
                    else:
                        st.error("Falha ao gravar ações. Verifique permissões/arquivo.")
        else:
            st.info("Nenhum equipamento com alerta dentro do threshold configurado.")

        st.markdown("---")
        st.subheader("Visão geral da frota - manutenção planejada")
        overview_cols = ["Cod_Equip", "DESCRICAO_EQUIPAMENTO", "Km_Hr_Atual", "Unid",
                         "Rev1_Next", "Rev1_To_Go", "Rev2_Next", "Rev2_To_Go", "Rev3_Next", "Rev3_To_Go"]
        available_over = [c for c in overview_cols if c in mf.columns]
        st.dataframe(mf[available_over].sort_values("Cod_Equip").reset_index(drop=True), use_container_width=True)
        st.download_button("⬇️ Exportar CSV - Plano de Manutenção (Visão Geral)", mf[available_over].to_csv(index=False).encode("utf-8"), "manutencao_overview.csv", "text/csv")

if __name__ == "__main__":
    main()
